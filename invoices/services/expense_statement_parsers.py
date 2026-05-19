from __future__ import annotations

import csv
import io
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

from invoices.models import normalize_import_header


class StatementParseError(Exception):
    """Raised when an uploaded statement file cannot be parsed deterministically."""


@dataclass
class ParsedStatementFile:
    source_name: str
    headers: list[str]
    rows: list[dict[str, str]]


def parse_statement_upload(raw_bytes: bytes, filename: str) -> list[ParsedStatementFile]:
    """Parse one uploaded statement, checking spreadsheets before generic ZIP files."""
    if _is_xlsx(filename, raw_bytes):
        return [_parse_xlsx(raw_bytes, filename)]
    if _is_xls(filename, raw_bytes):
        return [_parse_xls(raw_bytes, filename)]
    if filename.lower().endswith('.zip') or zipfile.is_zipfile(io.BytesIO(raw_bytes)):
        return _parse_zip(raw_bytes, filename)
    return [_parse_csv_bytes(raw_bytes, filename)]


def _is_xlsx(filename: str, raw_bytes: bytes) -> bool:
    return filename.lower().endswith(('.xlsx', '.xlsm'))


def _is_xls(filename: str, raw_bytes: bytes) -> bool:
    return filename.lower().endswith('.xls') or raw_bytes.startswith(b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1')


def _parse_zip(raw_bytes: bytes, filename: str) -> list[ParsedStatementFile]:
    try:
        archive = zipfile.ZipFile(io.BytesIO(raw_bytes))
    except zipfile.BadZipFile as exc:
        raise StatementParseError(f'{filename}: invalid ZIP archive.') from exc
    parsed = []
    for member in archive.infolist():
        if member.is_dir() or not member.filename.lower().endswith('.csv'):
            continue
        with archive.open(member, 'r') as handle:
            parsed.append(_parse_csv_bytes(handle.read(), member.filename))
    if not parsed:
        raise StatementParseError(f'{filename}: ZIP does not contain CSV files.')
    return parsed


def _parse_csv_bytes(raw_bytes: bytes, source_name: str) -> ParsedStatementFile:
    text = _decode(raw_bytes)
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    parsed = _rows_to_parsed_file(source_name, reader)
    if parsed is None:
        raise StatementParseError(f'{source_name}: missing CSV headers.')
    return parsed


def _parse_xlsx(raw_bytes: bytes, source_name: str) -> ParsedStatementFile:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised only in incomplete installs
        raise StatementParseError('XLSX imports require openpyxl to be installed.') from exc
    try:
        workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise StatementParseError(f'{source_name}: invalid or unsupported XLSX file.') from exc
    try:
        for worksheet in workbook.worksheets:
            if getattr(worksheet, 'sheet_state', 'visible') != 'visible':
                continue
            parsed = _rows_to_parsed_file(source_name, worksheet.iter_rows(values_only=True))
            if parsed is not None:
                return parsed
    finally:
        workbook.close()
    raise StatementParseError(f'{source_name}: no worksheet table found.')


def _parse_xls(raw_bytes: bytes, source_name: str) -> ParsedStatementFile:
    try:
        import xlrd
    except ImportError as exc:  # pragma: no cover - exercised only in incomplete installs
        raise StatementParseError('XLS imports require xlrd to be installed.') from exc
    try:
        workbook = xlrd.open_workbook(file_contents=raw_bytes)
    except Exception as exc:
        raise StatementParseError(f'{source_name}: invalid or unsupported XLS file.') from exc
    for sheet in workbook.sheets():
        if getattr(sheet, 'visibility', 0) != 0:
            continue
        datemode = getattr(workbook, 'datemode', 0)
        rows = (_xls_row_values(sheet, row_index, datemode) for row_index in range(sheet.nrows))
        parsed = _rows_to_parsed_file(source_name, rows)
        if parsed is not None:
            return parsed
    raise StatementParseError(f'{source_name}: no worksheet table found.')


def _xls_row_values(sheet: Any, row_index: int, datemode: int) -> list[Any]:
    try:
        import xlrd
    except ImportError:  # pragma: no cover
        xlrd = None
    values = []
    for cell in sheet.row(row_index):
        value = cell.value
        if xlrd is not None and cell.ctype == xlrd.XL_CELL_DATE:
            try:
                value = xlrd.xldate.xldate_as_datetime(value, datemode)
            except Exception:
                pass
        values.append(value)
    return values


def _rows_to_parsed_file(source_name: str, rows: Any) -> ParsedStatementFile | None:
    buffered_rows = [list(row) for row in rows]
    header_index = _detect_header_index(buffered_rows)
    if header_index is None:
        return None
    headers = [_stringify_cell(value) for value in buffered_rows[header_index]]
    last_header = max(index for index, header in enumerate(headers) if header.strip())
    headers = _dedupe_headers(headers[: last_header + 1])
    parsed_rows: list[dict[str, str]] = []
    for row in buffered_rows[header_index + 1 :]:
        values = [_stringify_cell(value) for value in row[: len(headers)]]
        values.extend([''] * (len(headers) - len(values)))
        if not any(value.strip() for value in values):
            continue
        if _looks_like_trailing_non_transaction(headers, values):
            continue
        parsed_rows.append(dict(zip(headers, values)))
    return ParsedStatementFile(source_name=source_name, headers=headers, rows=parsed_rows)


def _detect_header_index(rows: list[list[Any]]) -> int | None:
    for index, row in enumerate(rows):
        labels = [_stringify_cell(value) for value in row]
        non_blank = [label for label in labels if label.strip()]
        if len(non_blank) < 2:
            continue
        normalized = {normalize_import_header(label) for label in non_blank}
        if any(label in normalized for label in ('date', 'posted', 'amount', 'total', 'description', 'memo', 'item')):
            return index
        if len(non_blank) >= 3 and all(not _looks_like_number(label) for label in non_blank[:3]):
            return index
    return None


def _looks_like_trailing_non_transaction(headers: list[str], values: list[str]) -> bool:
    normalized_headers = [normalize_import_header(header) for header in headers]
    non_blank_count = sum(1 for value in values if value.strip())
    if non_blank_count <= 1:
        return True
    date_indexes = [index for index, header in enumerate(normalized_headers) if header in {'date', 'posted', 'value date'}]
    amount_indexes = [index for index, header in enumerate(normalized_headers) if header in {'amount', 'total'}]
    if date_indexes and amount_indexes:
        has_date = any(index < len(values) and _looks_like_date(values[index]) for index in date_indexes)
        if not has_date:
            return True
    return False


def _looks_like_date(value: str) -> bool:
    value = value.strip()
    if not value:
        return False
    return any(separator in value for separator in ('/', '-')) and any(character.isdigit() for character in value)


def _dedupe_headers(headers: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    deduped = []
    for index, header in enumerate(headers, start=1):
        label = header.strip() or f'Column {index}'
        count = counts.get(label, 0) + 1
        counts[label] = count
        deduped.append(label if count == 1 else f'{label} {count}')
    return deduped


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        if value.time() == time.min:
            return value.date().isoformat()
        return value.isoformat(sep=' ')
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, 'f')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_number(value: str) -> bool:
    try:
        Decimal(value.replace(',', ''))
    except Exception:
        return False
    return True


def _decode(raw_bytes: bytes) -> str:
    for encoding in ('utf-8-sig', 'utf-16', 'latin-1'):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode('utf-8-sig', errors='replace')
